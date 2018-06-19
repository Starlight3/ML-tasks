import PyPDF2
from rake import rake
import openpyxl

pdfFileObj = open('Java.pdf', 'rb')
pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
 
no_pages = pdfReader.numPages
s = ''
for i in range(0, no_pages):
    pageObj = pdfReader.getPage(i)
    s += pageObj.extractText()
    #print(rake(s))
    i+=1

pairs = rake(s)
wb = openpyxl.Workbook()
sheet = wb.active
cell = sheet.cell(row=1, column=1)
cell.value = "Keywords"
cell = sheet.cell(row=1, column=2)
cell.value = "Weights"
i=2
for pair in pairs:
    cell = sheet.cell(row=i, column=1)
    cell.value = pair[0]
    cell = sheet.cell(row=i, column=2)
    cell.value = pair[1]
    i+=1
wb.save('Java Keywords.xlsx')

pdfFileObj.close()
